#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
import subprocess
import time
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DOWNLOAD_DIR = ROOT / 'file'
SITE_URL = 'https://tcziryzf7x.jiandaoyun.com/dash/699fae71c3a3b520445a159f'
SITE_PASSWORD = '20260226@Zj'
DEBUG_LOG = ROOT / 'references' / 'debug-log.md'
CURL_USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36'
CURL_REFERER = 'https://tcziryzf7x.jiandaoyun.com/'
TABLE_STATE_JS = r'''
(() => {
  function titleText(title) {
    if (title == null) return '';
    if (typeof title === 'string' || typeof title === 'number') return String(title).trim();
    if (Array.isArray(title)) return title.map(titleText).join('').trim();
    if (title.props && title.props.children != null) return titleText(title.props.children);
    return '';
  }

  const pagerInput = document.querySelector('.x-pagination .page-input input.input-inner');
  const totalEl = document.querySelector('.x-pagination .total-page');
  const currentPage = pagerInput ? Number(pagerInput.value || 1) : 1;
  const totalPages = totalEl ? Number((totalEl.textContent.match(/\/\s*(\d+)/) || [])[1] || 1) : 1;

  const rows = Array.from(document.querySelectorAll('tr.table-row')).map((tr) => {
    const instKey = Object.keys(tr).find(k => k.startsWith('__reactInternalInstance$'));
    const inst = instKey ? tr[instKey] : null;
    const children = inst?.pendingProps?.children || inst?.memoizedProps?.children || [];
    const row = children?.[0]?.props?.row || {};
    const columns = [];

    for (const child of children) {
      const column = child?.props?.column || {};
      const title = titleText(column.title);
      const dataIndex = column.dataIndex || '';
      if (!title || !dataIndex) continue;
      columns.push({
        title,
        dataIndex,
        fieldType: column.fieldType || '',
        widgetType: column.widgetType || '',
        value: row[dataIndex] ?? null,
      });
    }

    return { columns };
  });

  return { currentPage, totalPages, rows };
})()
'''


def log_debug(message: str) -> None:
    DEBUG_LOG.parent.mkdir(parents=True, exist_ok=True)
    with open(DEBUG_LOG, 'a', encoding='utf-8') as f:
        f.write(f'\n- {message}')



def run(cmd: list[str], check: bool = True, capture: bool = True) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(cmd, text=True, capture_output=capture)
    if check and result.returncode != 0:
        cmd_str = ' '.join(str(x) for x in cmd)
        raise RuntimeError(
            f'Command failed ({result.returncode}): {cmd_str}\n'
            f'STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}'
        )
    return result



def normalize_date(v: Any) -> str:
    if pd.isna(v) or v == '':
        return ''
    try:
        dt = pd.to_datetime(v)
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return str(v).strip()



def normalize_amount(v: Any) -> str:
    if pd.isna(v) or v == '':
        return ''
    s = str(v).replace(',', '').strip()
    try:
        return format(Decimal(s).quantize(Decimal('0.01')), 'f')
    except (InvalidOperation, ValueError):
        return s



def normalize_text(v: Any) -> str:
    if v is None or pd.isna(v):
        return ''
    return re.sub(r'\s+', ' ', str(v).strip())



def normalize_voucher_no(v: Any) -> str:
    if v is None or (not isinstance(v, str) and pd.isna(v)):
        return ''
    s = str(v).strip()
    if re.fullmatch(r'-?\d+\.0+', s):
        return s.split('.', 1)[0]
    return s



def safe_name(s: Any) -> str:
    s = normalize_text(s)
    s = re.sub(r'[\\/:*?"<>|]+', '_', s)
    s = re.sub(r'\s+', ' ', s)
    return s[:180] or 'unnamed'



def extract_company_from_filename(path: Path) -> str:
    return path.stem.split('_', 1)[0].strip()



def ensure_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in ['下载附件名称', '下载状态', '下载文件数', '下载目录', '失败原因']:
        if col not in df.columns:
            df[col] = ''
    return df



def parse_excel(path: Path):
    df = pd.read_excel(path)
    df = ensure_output_columns(df)
    company = extract_company_from_filename(path)
    rows: list[dict[str, Any]] = []
    for idx, row in df.iterrows():
        voucher_no = normalize_voucher_no(row.get('凭证编号'))
        date_str = normalize_date(row.get('凭证日期'))
        amount = normalize_amount(row.get('借方发生金额'))
        if amount in ('', '0.00', '0'):
            amount = normalize_amount(row.get('贷方发生金额'))
        rows.append({
            'index': idx,
            '序号': row.get('序号', idx + 1),
            '公司名称': company,
            '字': normalize_text(row.get('字')),
            '凭证编号': voucher_no,
            '凭证日期': date_str,
            '摘要': normalize_text(row.get('摘要')),
            '金额': amount,
            '附件索引号': normalize_text(row.get('附件索引号')),
            '附件内容填写': normalize_text(row.get('附件内容填写')),
        })
    return df, rows, company



def agent_browser(*args: str, check: bool = True) -> subprocess.CompletedProcess[str]:
    cmd = ['npx', '-y', 'agent-browser', *args]
    return run(cmd, check=check, capture=True)



def ab_eval(js: str) -> str:
    proc = subprocess.run(['npx', '-y', 'agent-browser', 'eval', '--stdin'], input=js, text=True, capture_output=True)
    if proc.returncode != 0:
        raise RuntimeError(f'eval failed\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}')
    return proc.stdout.strip()



def parse_eval_output(raw: str) -> Any:
    raw = raw.strip()
    if raw == '':
        return None
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return raw
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith('{') or stripped.startswith('['):
            try:
                return json.loads(stripped)
            except json.JSONDecodeError:
                return value
    return value



def ab_eval_json(js: str) -> Any:
    return parse_eval_output(ab_eval(js))



def login_if_needed() -> str:
    agent_browser('open', SITE_URL)
    agent_browser('wait', '--load', 'networkidle')
    snap = agent_browser('snapshot', '-i').stdout
    if 'Please enter your password' in snap or 'password' in snap.lower():
        js = f'''
(() => {{
  const input = document.querySelector('input[type="password"], input');
  if (!input) return 'NO_INPUT';
  input.focus();
  input.value = {json.dumps(SITE_PASSWORD)};
  input.dispatchEvent(new Event('input', {{bubbles:true}}));
  input.dispatchEvent(new Event('change', {{bubbles:true}}));
  const btn = Array.from(document.querySelectorAll('button')).find(el => !el.disabled);
  if (!btn) return 'NO_BUTTON';
  btn.click();
  return 'OK';
}})()
'''
        result = ab_eval(js)
        log_debug(f'login_if_needed password_submit={result}')
        agent_browser('wait', '--load', 'networkidle')
    return agent_browser('get', 'title').stdout.strip()



def close_dialog_if_any() -> None:
    try:
        ab_eval(
            """
(() => {
  const btn = Array.from(document.querySelectorAll('button, .icon-close, .close, .x-icon')).find(el => {
    const t = (el.innerText || el.getAttribute('aria-label') || el.getAttribute('title') || '');
    return /关闭|close|×|✕/.test(t);
  });
  if (btn) { btn.click(); return 'CLOSED'; }
  return 'NO_CLOSE';
})()
"""
        )
    except Exception:
        pass



def set_page_number(page: int) -> None:
    for _attempt in range(3):
        js = f'''
(() => {{
  const input = document.querySelector('.x-pagination .page-input input.input-inner');
  if (!input) return 'NO_PAGE_INPUT';
  const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
  input.focus();
  setter.call(input, {json.dumps(str(page))});
  input.dispatchEvent(new InputEvent('input', {{ bubbles: true, data: {json.dumps(str(page))}, inputType: 'insertText' }}));
  input.dispatchEvent(new Event('change', {{ bubbles: true }}));
  input.dispatchEvent(new KeyboardEvent('keydown', {{ bubbles: true, key: 'Enter', code: 'Enter' }}));
  input.dispatchEvent(new KeyboardEvent('keyup', {{ bubbles: true, key: 'Enter', code: 'Enter' }}));
  input.blur();
  return input.value;
}})()
'''
        result = ab_eval(js)
        log_debug(f'set_page_number page={page} result={result}')
        if result != 'NO_PAGE_INPUT':
            for _ in range(40):
                time.sleep(0.3)
                state = extract_current_page_state()
                if state['currentPage'] == page:
                    return
        time.sleep(0.5)
    raise RuntimeError(f'failed to jump to page {page}')



def wait_for_month_filter(target_month: str, timeout_seconds: float = 20.0) -> None:
    target_year, target_mon = target_month.split('-')
    deadline = time.time() + timeout_seconds
    last_inputs = None
    last_row_count = None
    while time.time() < deadline:
        inputs = ab_eval_json(
            """
(() => Array.from(document.querySelectorAll('input.input-inner')).filter(el => el.offsetParent).slice(0, 2).map(x => x.value))()
"""
        )
        state = extract_current_page_state()
        rows = state['rows']
        months_ok = bool(rows)
        for row in rows:
            if row['年'] and row['月']:
                if not (row['年'] == target_year and row['月'] == target_mon):
                    months_ok = False
                    break
        if inputs == [target_month, target_month] and months_ok:
            log_debug(f'wait_for_month_filter stabilized month={target_month} rows={len(rows)} page={state["currentPage"]}/{state["totalPages"]}')
            return
        last_inputs = inputs
        last_row_count = len(rows)
        time.sleep(0.5)
    raise RuntimeError(
        f'month filter did not stabilize for {target_month}; last_inputs={last_inputs} last_row_count={last_row_count}'
    )



def set_month_range(start_month: str | None, end_month: str | None) -> None:
    if not start_month or not end_month:
        return
    js = f'''
(() => {{
  const inputs = Array.from(document.querySelectorAll('input.input-inner')).filter(el => el.offsetParent).slice(0, 2);
  if (inputs.length < 2) return 'NO_DATE_INPUTS';
  const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
  function apply(input, value) {{
    input.focus();
    setter.call(input, '');
    input.dispatchEvent(new InputEvent('input', {{ bubbles: true, data: null, inputType: 'deleteContentBackward' }}));
    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
    setter.call(input, value);
    input.dispatchEvent(new InputEvent('input', {{ bubbles: true, data: value, inputType: 'insertText' }}));
    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
    input.dispatchEvent(new KeyboardEvent('keydown', {{ bubbles: true, key: 'Enter', code: 'Enter' }}));
    input.dispatchEvent(new KeyboardEvent('keyup', {{ bubbles: true, key: 'Enter', code: 'Enter' }}));
    input.blur();
  }}
  // 用户要求：先结束日期，再开始日期
  apply(inputs[1], {json.dumps(end_month)});
  apply(inputs[0], {json.dumps(start_month)});
  return JSON.stringify(inputs.map(x => x.value));
}})()
'''
    last_error = None
    for attempt in range(1, 4):
        result = ab_eval(js)
        log_debug(f'set_month_range attempt={attempt} {start_month}..{end_month} => {result}')
        time.sleep(1.5)
        try:
            wait_for_month_filter(start_month, timeout_seconds=8.0)
            set_page_number(1)
            return
        except Exception as e:
            last_error = e
            log_debug(f'set_month_range retry attempt={attempt} failed error={e}')
            close_dialog_if_any()
            time.sleep(1.0)
    raise RuntimeError(f'failed to set month range {start_month}..{end_month}: {last_error}')



def extract_current_page_state() -> dict[str, Any]:
    state = ab_eval_json(TABLE_STATE_JS)
    if not isinstance(state, dict):
        raise RuntimeError(f'unexpected table state: {state!r}')
    rows = []
    for item in state.get('rows', []):
        columns = item.get('columns', []) or []
        colmap: dict[str, Any] = {}
        for col in columns:
            title = normalize_text(col.get('title'))
            if title:
                colmap[title] = col.get('value')
        year = normalize_voucher_no(colmap.get('年'))
        month_raw = normalize_voucher_no(colmap.get('月'))
        month = month_raw.zfill(2) if month_raw else ''
        rows.append({
            '公司名称': normalize_text(colmap.get('公司名称')),
            '年': year,
            '月': month,
            '凭证字': normalize_text(colmap.get('凭证字')),
            '凭证号': normalize_voucher_no(colmap.get('凭证号')),
            '摘要': normalize_text(colmap.get('摘要')),
            '金额(元)': normalize_amount(colmap.get('金额(元)')),
            '会计凭证': colmap.get('会计凭证') or [],
            '银行回单': colmap.get('银行回单') or [],
            '审批单': colmap.get('审批单') or [],
            '财务档案': colmap.get('财务档案') or [],
            '凭证单据编号': normalize_text(colmap.get('凭证单据编号')),
            '_raw': colmap,
        })
    return {
        'currentPage': int(state.get('currentPage') or 1),
        'totalPages': int(state.get('totalPages') or 1),
        'rows': rows,
    }



def click_next_page(prev_page: int) -> bool:
    result = ab_eval(
        """
(() => {
  const btn = document.querySelector('.page-turn-next');
  if (!btn || btn.classList.contains('page-btn-disabled')) return 'NO_NEXT';
  btn.click();
  return 'OK';
})()
"""
    )
    if 'NO_NEXT' in str(result):
        return False
    for _ in range(40):
        time.sleep(0.4)
        state = extract_current_page_state()
        if state['currentPage'] != prev_page:
            return True
    raise RuntimeError(f'page did not advance from {prev_page}')



def crawl_site_rows(max_pages: int | None = None) -> list[dict[str, Any]]:
    seen_pages: set[int] = set()
    all_rows: list[dict[str, Any]] = []
    while True:
        state = extract_current_page_state()
        page = state['currentPage']
        total = state['totalPages']
        if page in seen_pages:
            break
        seen_pages.add(page)
        all_rows.extend(state['rows'])
        log_debug(f'crawl_site_rows page={page}/{total} rows={len(state["rows"])} total_acc={len(all_rows)}')
        if max_pages and len(seen_pages) >= max_pages:
            break
        if page >= total:
            break
        if not click_next_page(page):
            break
    return all_rows



def crawl_site_rows_for_months(months: list[str], max_pages_per_month: int | None = None) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    for month in months:
        close_dialog_if_any()
        set_month_range(month, month)
        month_rows = crawl_site_rows(max_pages=max_pages_per_month)
        log_debug(f'crawl_site_rows_for_months month={month} rows={len(month_rows)}')
        all_rows.extend(month_rows)
    return dedupe_site_rows(all_rows)



def dedupe_site_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        voucher_files = tuple(sorted((f.get('downloadUrl') or '', f.get('name') or '') for f in row.get('会计凭证', []) if isinstance(f, dict)))
        key = (
            row.get('公司名称', ''),
            row.get('年', ''),
            row.get('月', ''),
            row.get('凭证号', ''),
            row.get('摘要', ''),
            row.get('金额(元)', ''),
            row.get('凭证单据编号', ''),
            voucher_files,
        )
        deduped[key] = row
    return list(deduped.values())



def amounts_equal(a: str, b: str) -> bool:
    if not a or not b:
        return False
    try:
        return Decimal(a) == Decimal(b)
    except Exception:
        return a == b



def summary_matches(excel_summary: str, site_summary: str) -> bool:
    a = normalize_text(excel_summary)
    b = normalize_text(site_summary)
    if not a or not b:
        return False
    return a == b or a in b or b in a



def build_indexes(site_rows: list[dict[str, Any]]):
    by_month_voucher: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    by_attachment_no: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in site_rows:
        key = (
            normalize_text(row.get('公司名称')),
            normalize_text(row.get('年')),
            normalize_text(row.get('月')),
            normalize_text(row.get('凭证号')),
        )
        by_month_voucher[key].append(row)
        attachment_no = normalize_text(row.get('凭证单据编号'))
        if attachment_no:
            by_attachment_no[attachment_no].append(row)
    return by_month_voucher, by_attachment_no



def choose_best_candidate(excel_row: dict[str, Any], candidates: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, dict[str, Any]]:
    if not candidates:
        return None, {'reason': 'no-candidate'}

    scored: list[tuple[int, dict[str, Any]]] = []
    for cand in candidates:
        score = 0
        if cand.get('会计凭证'):
            score += 100
        if excel_row.get('附件索引号') and excel_row['附件索引号'] == normalize_text(cand.get('凭证单据编号')):
            score += 80
        if excel_row.get('金额') and amounts_equal(excel_row['金额'], normalize_amount(cand.get('金额(元)'))):
            score += 20
        if excel_row.get('摘要') and summary_matches(excel_row['摘要'], normalize_text(cand.get('摘要'))):
            score += 10
        if excel_row.get('字') and excel_row['字'] == normalize_text(cand.get('凭证字')):
            score += 5
        scored.append((score, cand))

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best = scored[0]
    ambiguous = len(scored) > 1 and scored[1][0] == best_score
    meta = {
        'reason': 'matched',
        'candidate_count': len(candidates),
        'best_score': best_score,
        'ambiguous': ambiguous,
    }
    return best, meta



def match_site_row(excel_row: dict[str, Any], by_month_voucher, by_attachment_no) -> tuple[dict[str, Any] | None, dict[str, Any]]:
    if not excel_row.get('凭证日期') or not excel_row.get('凭证编号'):
        return None, {'reason': 'missing-date-or-voucher'}

    try:
        year, month, _day = excel_row['凭证日期'].split('-')
    except ValueError:
        return None, {'reason': 'bad-date-format'}

    base_key = (excel_row['公司名称'], year, month, excel_row['凭证编号'])
    candidates = list(by_month_voucher.get(base_key, []))

    attachment_no = excel_row.get('附件索引号')
    if attachment_no:
        attachment_candidates = by_attachment_no.get(attachment_no, [])
        if candidates:
            filtered = [c for c in candidates if normalize_text(c.get('凭证单据编号')) == attachment_no]
            if filtered:
                candidates = filtered
        elif attachment_candidates:
            candidates = attachment_candidates

    best, meta = choose_best_candidate(excel_row, candidates)
    if best is None:
        meta['lookup_key'] = base_key
    return best, meta



def filename_from_url(url: str, fallback: str | None = None) -> str:
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    if 'attname' in qs and qs['attname']:
        return safe_name(unquote(qs['attname'][0]))
    if fallback:
        return safe_name(fallback)
    return safe_name(Path(parsed.path).name or f'download-{int(time.time())}.bin')



def download_file(url: str, out_dir: Path, filename_hint: str | None = None) -> tuple[str, str]:
    if not isinstance(url, str) or not url:
        raise ValueError(f'invalid download url: {url!r}')
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = filename_from_url(url, filename_hint)
    dest = out_dir / filename
    stem = dest.stem
    suffix = dest.suffix
    n = 2

    while dest.exists():
        dest = out_dir / f'{stem}_{n}{suffix}'
        n += 1

    part = dest.with_suffix(dest.suffix + '.part')
    if part.exists():
        part.unlink()

    cmd = [
        'curl',
        '-fL',
        '--retry', '3',
        '--retry-delay', '2',
        '--connect-timeout', '30',
        '--max-time', '180',
        '-A', CURL_USER_AGENT,
        '-e', CURL_REFERER,
        '-H', 'Accept: */*',
        '-o', str(part),
        url,
    ]
    result = run(cmd, check=False, capture=True)
    if result.returncode != 0:
        raise RuntimeError(f'curl failed rc={result.returncode} stderr={result.stderr.strip()} stdout={result.stdout.strip()}')
    if not part.exists() or part.stat().st_size == 0:
        raise RuntimeError('curl finished but no file was written')
    part.rename(dest)
    return dest.name, str(dest)



def prepare_run_dir(download_root: Path, company: str) -> Path:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    run_dir = download_root / safe_name(company) / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir



def build_output_paths(excel_path: Path, run_dir: Path, output_excel_arg: str | None):
    default_name = f'{safe_name(excel_path.stem)}_result.xlsx'
    output_excel = Path(output_excel_arg).resolve() if output_excel_arg else (run_dir / default_name)
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    stem_no_suffix = output_excel.stem
    csv_path = output_excel.with_name(f'{stem_no_suffix}.csv')
    json_path = output_excel.with_name(f'{stem_no_suffix}.json')
    return output_excel, csv_path, json_path



def download_accounting_vouchers(site_row: dict[str, Any], run_dir: Path, cache: dict[str, dict[str, str]]):
    downloaded = []
    for item in site_row.get('会计凭证', []):
        if not isinstance(item, dict):
            continue
        url = item.get('downloadUrl') or ''
        name_hint = item.get('name') or ''
        if not url:
            continue
        if url in cache and Path(cache[url]['path']).exists():
            downloaded.append({'name': cache[url]['name'], 'path': cache[url]['path'], 'url': url, 'cached': True})
            continue
        name, path = download_file(url, run_dir, name_hint)
        cache[url] = {'name': name, 'path': path}
        downloaded.append({'name': name, 'path': path, 'url': url, 'cached': False})
    return downloaded, str(run_dir)



def process_row(excel_row: dict[str, Any], by_month_voucher, by_attachment_no, run_dir: Path, cache: dict[str, dict[str, str]]):
    site_row, match_meta = match_site_row(excel_row, by_month_voucher, by_attachment_no)
    if not site_row:
        return {
            'matched': False,
            'downloaded': [],
            'download_dir': '',
            'match_meta': match_meta,
            'site_row': None,
        }
    if not site_row.get('会计凭证'):
        return {
            'matched': True,
            'downloaded': [],
            'download_dir': '',
            'match_meta': {**match_meta, 'reason': 'matched-but-no-accounting-voucher'},
            'site_row': {
                '公司名称': site_row.get('公司名称'),
                '年': site_row.get('年'),
                '月': site_row.get('月'),
                '凭证号': site_row.get('凭证号'),
                '凭证单据编号': site_row.get('凭证单据编号'),
                '摘要': site_row.get('摘要'),
            },
        }

    downloaded, download_dir = download_accounting_vouchers(site_row, run_dir, cache)
    return {
        'matched': True,
        'downloaded': downloaded,
        'download_dir': download_dir,
        'match_meta': match_meta,
        'site_row': {
            '公司名称': site_row.get('公司名称'),
            '年': site_row.get('年'),
            '月': site_row.get('月'),
            '凭证号': site_row.get('凭证号'),
            '凭证单据编号': site_row.get('凭证单据编号'),
            '摘要': site_row.get('摘要'),
            '金额(元)': site_row.get('金额(元)'),
            '会计凭证文件数': len(site_row.get('会计凭证', [])),
        },
    }



def summarize_downloaded(downloaded: list[dict[str, Any]] | None) -> tuple[list[str], list[str]]:
    names: list[str] = []
    paths: list[str] = []
    for item in downloaded or []:
        if not isinstance(item, dict):
            continue
        name = normalize_text(item.get('name'))
        path = str(item.get('path') or '').strip()
        if not name or not path:
            continue
        if not Path(path).exists():
            continue
        names.append(name)
        paths.append(path)
    return names, paths



def update_row_output(df: pd.DataFrame, row_index: int, result: dict[str, Any]) -> tuple[list[str], list[str]]:
    names, paths = summarize_downloaded(result.get('downloaded'))
    df.at[row_index, '下载附件名称'] = '\n'.join(names)
    df.at[row_index, '下载状态'] = 'success' if names else ('matched-no-download' if result.get('matched') else 'not-found')
    df.at[row_index, '下载文件数'] = len(names)
    df.at[row_index, '下载目录'] = result.get('download_dir', '')
    df.at[row_index, '失败原因'] = '' if names else result.get('match_meta', {}).get('reason', '')
    return names, paths



def build_result_record(row: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    names, _paths = summarize_downloaded(result.get('downloaded'))
    status = 'ok' if names else 'no-download'
    return {**row, **result, 'status': status}



def find_retry_rows(rows_to_process: list[dict[str, Any]], df: pd.DataFrame, results_by_index: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    retry_rows: list[dict[str, Any]] = []
    for row in rows_to_process:
        idx = row['index']
        result = results_by_index.get(idx) or {}
        names, paths = summarize_downloaded(result.get('downloaded'))
        cell_value = normalize_text(df.at[idx, '下载附件名称']) if '下载附件名称' in df.columns else ''
        status = normalize_text(df.at[idx, '下载状态']) if '下载状态' in df.columns else ''
        if status == 'success' and cell_value and names and all(Path(p).exists() for p in paths):
            continue
        retry_rows.append(row)
    return retry_rows



def retry_failed_rows(
    rows_to_process: list[dict[str, Any]],
    df: pd.DataFrame,
    results_by_index: dict[int, dict[str, Any]],
    hyperlink_map: dict[int, dict[str, Any]],
    run_dir: Path,
    download_cache: dict[str, dict[str, str]],
    max_pages_per_month: int | None = None,
    retry_rounds: int = 1,
) -> list[dict[str, Any]]:
    rounds: list[dict[str, Any]] = []
    for attempt in range(1, retry_rounds + 1):
        retry_rows = find_retry_rows(rows_to_process, df, results_by_index)
        if not retry_rows:
            break
        retry_months = month_list_from_rows(retry_rows)
        log_debug(f'retry_failed_rows attempt={attempt} rows={len(retry_rows)} months={retry_months}')
        site_rows = crawl_site_rows_for_months(retry_months, max_pages_per_month=max_pages_per_month)
        by_month_voucher, by_attachment_no = build_indexes(site_rows)
        recovered = 0
        for row in retry_rows:
            try:
                result = process_row(row, by_month_voucher, by_attachment_no, run_dir, download_cache)
                names, paths = update_row_output(df, row['index'], result)
                results_by_index[row['index']] = build_result_record(row, result)
                if names:
                    hyperlink_map[row['index']] = {
                        'names': names,
                        'paths': paths,
                        'dir_path': result.get('download_dir', ''),
                    }
                    recovered += 1
                else:
                    hyperlink_map.pop(row['index'], None)
            except Exception as e:
                df.at[row['index'], '下载附件名称'] = ''
                df.at[row['index'], '下载状态'] = 'error'
                df.at[row['index'], '失败原因'] = str(e)
                hyperlink_map.pop(row['index'], None)
                results_by_index[row['index']] = {**row, 'status': 'error', 'error': str(e), 'downloaded': []}
            finally:
                close_dialog_if_any()
        round_info = {
            'attempt': attempt,
            'retry_rows': len(retry_rows),
            'recovered': recovered,
            'months': retry_months,
            'site_rows': len(site_rows),
        }
        rounds.append(round_info)
        log_debug(f'retry_failed_rows_summary {round_info}')
        if recovered == 0:
            break
    return rounds



def month_list_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    return sorted({row['凭证日期'][:7] for row in rows if row.get('凭证日期') and len(row['凭证日期']) >= 7})



def apply_excel_hyperlinks(output_excel: Path, hyperlink_map: dict[int, dict[str, Any]]) -> None:
    wb = load_workbook(output_excel)
    ws = wb.active
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    name_col = header_map.get('下载附件名称')
    dir_col = header_map.get('下载目录')
    if not name_col:
        wb.save(output_excel)
        return

    for df_index, item in hyperlink_map.items():
        excel_row = df_index + 2
        cell = ws.cell(row=excel_row, column=name_col)
        names = item.get('names', [])
        paths = item.get('paths', [])
        dir_path = item.get('dir_path', '')
        if not names:
            continue
        cell.value = '\n'.join(names)
        target = ''
        if len(paths) == 1:
            target = os.path.relpath(paths[0], output_excel.parent)
        elif len(paths) > 1 and dir_path:
            target = os.path.relpath(dir_path, output_excel.parent)
        if target:
            cell.hyperlink = target
            cell.style = 'Hyperlink'
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        if dir_col and dir_path:
            dir_cell = ws.cell(row=excel_row, column=dir_col)
            dir_cell.value = os.path.relpath(dir_path, output_excel.parent)
            dir_cell.hyperlink = dir_cell.value
            dir_cell.style = 'Hyperlink'

    wb.save(output_excel)



def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('excel', help='xlsx path')
    ap.add_argument('--limit', type=int, default=0, help='0 means all rows')
    ap.add_argument('--download-dir', default=str(DEFAULT_DOWNLOAD_DIR))
    ap.add_argument('--output-excel', default='', help='optional explicit xlsx output path')
    ap.add_argument('--max-pages', type=int, default=0, help='0 means crawl all pages for each month')
    args = ap.parse_args()

    excel_path = Path(args.excel).resolve()
    download_dir = Path(args.download_dir).resolve()

    df, rows, company = parse_excel(excel_path)
    rows_to_process = rows[: args.limit] if args.limit and args.limit > 0 else rows
    run_dir = prepare_run_dir(download_dir, company)
    output_excel, csv_path, json_path = build_output_paths(excel_path, run_dir, args.output_excel or None)

    title = login_if_needed()
    months = month_list_from_rows(rows_to_process)
    site_rows = crawl_site_rows_for_months(months, max_pages_per_month=args.max_pages or None)
    by_month_voucher, by_attachment_no = build_indexes(site_rows)
    log_debug(f'site_rows_deduped={len(site_rows)} keys_month_voucher={len(by_month_voucher)} keys_attachment_no={len(by_attachment_no)} months={months}')

    results_by_index: dict[int, dict[str, Any]] = {}
    hyperlink_map: dict[int, dict[str, Any]] = {}
    download_cache: dict[str, dict[str, str]] = {}

    for row in rows_to_process:
        try:
            result = process_row(row, by_month_voucher, by_attachment_no, run_dir, download_cache)
            names, paths = update_row_output(df, row['index'], result)
            if names:
                hyperlink_map[row['index']] = {
                    'names': names,
                    'paths': paths,
                    'dir_path': result.get('download_dir', ''),
                }
            results_by_index[row['index']] = build_result_record(row, result)
        except Exception as e:
            df.at[row['index'], '下载附件名称'] = ''
            df.at[row['index'], '下载状态'] = 'error'
            df.at[row['index'], '下载文件数'] = 0
            df.at[row['index'], '下载目录'] = ''
            df.at[row['index'], '失败原因'] = str(e)
            results_by_index[row['index']] = {**row, 'status': 'error', 'error': str(e), 'downloaded': []}
        finally:
            close_dialog_if_any()

    retry_rounds = retry_failed_rows(
        rows_to_process=rows_to_process,
        df=df,
        results_by_index=results_by_index,
        hyperlink_map=hyperlink_map,
        run_dir=run_dir,
        download_cache=download_cache,
        max_pages_per_month=args.max_pages or None,
        retry_rounds=2,
    )

    ordered_results = [results_by_index[row['index']] for row in rows_to_process]

    df.to_excel(output_excel, index=False)
    apply_excel_hyperlinks(output_excel, hyperlink_map)

    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=sorted({k for r in ordered_results for k in r.keys()}))
        writer.writeheader()
        writer.writerows(ordered_results)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'title': title,
            'company': company,
            'run_dir': str(run_dir),
            'crawled_site_rows': len(site_rows),
            'processed': len(ordered_results),
            'retry_rounds': retry_rounds,
            'results': ordered_results,
        }, f, ensure_ascii=False, indent=2)

    success_count = sum(1 for r in ordered_results if r.get('status') == 'ok')
    print(json.dumps({
        'title': title,
        'company': company,
        'run_dir': str(run_dir),
        'crawled_site_rows': len(site_rows),
        'processed': len(ordered_results),
        'success_count': success_count,
        'retry_rounds': retry_rounds,
        'output_excel': str(output_excel),
        'csv': str(csv_path),
        'json': str(json_path),
        'download_dir': str(run_dir),
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
